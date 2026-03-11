from __future__ import annotations

import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver import ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


# ============================================================
# CONFIG
# ============================================================

LCR_BASE = "https://lcr.churchofjesuschrist.org"
MEMBER_LIST_PAGE_URL = f"{LCR_BASE}/records/member-list?lang=eng"
MEMBER_LIST_API_URL = f"{LCR_BASE}/api/umlu/report/member-list?lang=eng&unitNumber={{unit_number}}"

UNIT_NUMBER = os.getenv("UNIT_NUMBER", "253022").strip()

USERNAME = os.getenv("LCR_USERNAME", "").strip()
PASSWORD = os.getenv("LCR_PASSWORD", "").strip()

OUTPUT_DIR = "data"

HEADLESS = True
LONG_WAIT = 60


# ============================================================
# HELPERS
# ============================================================

def log(msg: str) -> None:
    print(f"[INFO] {msg}")


def err(msg: str) -> None:
    print(f"[ERROR] {msg}", file=sys.stderr)


def ensure_dir(path: str | Path) -> Path:
    p = Path(path)
    p.mkdir(parents=True, exist_ok=True)
    return p


def best_name(person: dict) -> str:
    # Prefer the same list-style name you showed from the API
    candidates = [
        person.get("nameListPreferredLocal"),
        (person.get("nameFormats") or {}).get("listPreferredLocal"),
        person.get("householdNameDirectoryLocal"),
        person.get("houseHoldMemberNameForList"),
        person.get("nameGivenPreferredLocal"),
    ]
    for value in candidates:
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


# ============================================================
# SELENIUM LOGIN
# ============================================================

def make_driver() -> webdriver.Chrome:
    opts = ChromeOptions()
    if HEADLESS or os.getenv("CI", "").lower() == "true":
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1600,2200")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--lang=en-US")
    return webdriver.Chrome(options=opts)


def login(driver: webdriver.Chrome) -> None:
    if not USERNAME or not PASSWORD:
        err("Missing env vars LCR_USERNAME and/or LCR_PASSWORD")
        sys.exit(1)

    log("Opening LCR base login page")
    driver.get(LCR_BASE)

    try:
        user_input = WebDriverWait(driver, LONG_WAIT).until(
            EC.presence_of_element_located((By.ID, "username-input"))
        )
        user_input.clear()
        user_input.send_keys(USERNAME)
        user_input.send_keys(Keys.ENTER)

        pwd_input = WebDriverWait(driver, LONG_WAIT).until(
            EC.presence_of_element_located((By.ID, "password-input"))
        )
        pwd_input.clear()
        pwd_input.send_keys(PASSWORD)
        pwd_input.send_keys(Keys.ENTER)

        WebDriverWait(driver, LONG_WAIT).until(EC.url_contains(LCR_BASE))
        log("Login submitted successfully")
    except Exception as ex:
        raise RuntimeError("Automated login failed with the known username/password field flow.") from ex


def build_requests_session_from_driver(driver: webdriver.Chrome) -> requests.Session:
    session = requests.Session()

    for cookie in driver.get_cookies():
        session.cookies.set(
            cookie["name"],
            cookie["value"],
            domain=cookie.get("domain"),
            path=cookie.get("path", "/"),
        )

    session.headers.update(
        {
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/json, text/plain, */*",
            "Referer": MEMBER_LIST_PAGE_URL,
        }
    )
    return session


# ============================================================
# API + FILTERING
# ============================================================

def fetch_member_list(session: requests.Session, unit_number: str) -> List[dict]:
    url = MEMBER_LIST_API_URL.format(unit_number=unit_number)
    log(f"Fetching member roster: {url}")
    response = session.get(url, timeout=60)
    response.raise_for_status()

    payload = response.json()
    if not isinstance(payload, list):
        raise RuntimeError("Member-list API did not return a list.")

    return payload


def filter_prospective_elders(members: List[dict]) -> List[Dict[str, str]]:
    output: List[Dict[str, str]] = []

    for person in members:
        status_flags = person.get("personStatusFlags") or {}
        is_prospective = (
            person.get("isProspectiveElder") is True
            or person.get("prospectiveElder") is True
            or status_flags.get("prospectiveElder") is True
        )

        if not is_prospective:
            continue

        name = best_name(person)
        if not name:
            continue

        output.append(
            {
                "Name": name,
                "prospective Elder?": "yes",
                "assigned to minister?": "no",
                "Is assigned ministers?": "no",
            }
        )

    output.sort(key=lambda r: r["Name"].casefold())
    return output


# ============================================================
# EXCEL OUTPUT
# ============================================================

def write_excel(rows: List[Dict[str, str]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospective Elders"

    headers = [
        "Name",
        "prospective Elder?",
        "assigned to minister?",
        "Is assigned ministers?",
    ]

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    yes_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    no_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    bold = Font(bold=True)
    left = Alignment(horizontal="left", vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(headers)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = left if col_idx == 1 else center

    for row_idx, row_data in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row_data["Name"])
        ws.cell(row=row_idx, column=2, value=row_data["prospective Elder?"])
        ws.cell(row=row_idx, column=3, value=row_data["assigned to minister?"])
        ws.cell(row=row_idx, column=4, value=row_data["Is assigned ministers?"])

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 23

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = left if cell.column == 1 else center

            if cell.column in (2, 3, 4):
                value = str(cell.value).strip().lower() if cell.value is not None else ""
                if value == "yes":
                    cell.fill = yes_fill
                elif value == "no":
                    cell.fill = no_fill

    wb.save(out_path)


# ============================================================
# MAIN
# ============================================================

def main() -> int:
    if not USERNAME or not PASSWORD:
        err("Missing LCR_USERNAME and/or LCR_PASSWORD environment variables.")
        return 1

    output_dir = ensure_dir(OUTPUT_DIR)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = output_dir / f"prospective_elders_{timestamp}.xlsx"

    driver = make_driver()
    try:
        login(driver)

        # Touch the member list page once so the session is warm
        driver.get(MEMBER_LIST_PAGE_URL)

        session = build_requests_session_from_driver(driver)
        members = fetch_member_list(session, UNIT_NUMBER)
        log(f"Members returned from API: {len(members)}")

        rows = filter_prospective_elders(members)
        log(f"Prospective elders found: {len(rows)}")

        write_excel(rows, out_path)
        log(f"Excel output written to {out_path}")
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    return 0


if __name__ == "__main__":
    sys.exit(main())
