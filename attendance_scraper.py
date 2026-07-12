from __future__ import annotations

import json
import os
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Set, Tuple

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver import ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


# ============================================================
# CONFIG
# ============================================================

LCR_BASE = "https://lcr.churchofjesuschrist.org"
ATTENDANCE_PAGE_URL = f"{LCR_BASE}/report/class-and-quorum-attendance/overview?lang=eng"
# FIX: LCR moved the rendered member list under /mlt/.
MEMBER_LIST_PAGE_URL = f"{LCR_BASE}/mlt/records/member-list?lang=eng"

UNIT_NUMBER = os.getenv("UNIT_NUMBER", "253022").strip()

USERNAME = os.getenv("LCR_USERNAME", "").strip()
PASSWORD = os.getenv("LCR_PASSWORD", "").strip()

START_DATE = os.getenv("START_DATE", "2025-12-28").strip()
END_DATE = os.getenv("END_DATE", "2026-03-08").strip()

OUTPUT_DIR = "data"

HEADLESS = True
DEFAULT_WAIT = 30
LONG_WAIT = 60
ROSTER_WAIT = 180
WINDOW_DAYS = 28

DEBUG_MEMBER_HTML = Path(OUTPUT_DIR) / "debug_attendance_member_list.html"
DEBUG_MEMBER_TEXT = Path(OUTPUT_DIR) / "debug_attendance_member_list.txt"
DEBUG_MEMBER_ROWS = Path(OUTPUT_DIR) / "debug_attendance_member_rows.txt"
DEBUG_ATTENDANCE_JSON = Path(OUTPUT_DIR) / "debug_attendance_last_response.json"


# ============================================================
# HELPERS
# ============================================================

def log(msg: str) -> None:
    print(f"[INFO] {msg}")


def err(msg: str) -> None:
    print(f"[ERROR] {msg}", file=sys.stderr)


def ensure_dir(path: str | Path) -> Path:
    p = Path(path)
    p.mkdir(parents=True, exist_ok=True)
    return p


def parse_iso_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def to_iso(dt: date) -> str:
    return dt.strftime("%Y-%m-%d")


def format_excel_header(dt: date) -> str:
    return f"{dt.strftime('%b')} {dt.day} {dt.year}"


def sunday_on_or_before(dt: date) -> date:
    return dt - timedelta(days=(dt.weekday() + 1) % 7)


def build_date_windows(start_dt: date, end_dt: date) -> List[Tuple[date, date]]:
    windows: List[Tuple[date, date]] = []
    first_start = sunday_on_or_before(start_dt - timedelta(days=28))
    last_start = sunday_on_or_before(end_dt)

    current = first_start
    while current <= last_start:
        windows.append((current, current + timedelta(days=28)))
        current += timedelta(days=7)

    return windows


def get_body_text(driver: webdriver.Chrome) -> str:
    return driver.find_element(By.TAG_NAME, "body").text


def clean_name(name: str) -> str:
    name = re.sub(r"\s*(Out-of-Unit|Not Baptized)\s*", " ", name)
    name = re.sub(r"\s+", " ", name)
    return name.strip()


def looks_like_name(name: str) -> bool:
    if not name or name == "Come, Follow Me" or len(name) > 90:
        return False
    return bool(re.match(r"^[A-Za-zÀ-ÿ'’.\- ]+,\s+[A-Za-zÀ-ÿ'’.\- ]+$", name))


def first_string(obj: dict, keys: tuple[str, ...]) -> str:
    for key in keys:
        value = obj.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


# ============================================================
# SELENIUM LOGIN
# ============================================================

def make_driver() -> webdriver.Chrome:
    opts = ChromeOptions()
    if HEADLESS or os.getenv("CI", "").lower() == "true":
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1600,2200")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--lang=en-US")
    return webdriver.Chrome(options=opts)


def login(driver: webdriver.Chrome) -> None:
    if not USERNAME or not PASSWORD:
        raise RuntimeError("Missing LCR_USERNAME and/or LCR_PASSWORD.")

    log("Opening LCR login page")
    driver.get(LCR_BASE)

    user_input = WebDriverWait(driver, LONG_WAIT).until(
        EC.presence_of_element_located((By.ID, "username-input"))
    )
    user_input.clear()
    user_input.send_keys(USERNAME)
    user_input.send_keys(Keys.ENTER)

    pwd_input = WebDriverWait(driver, LONG_WAIT).until(
        EC.presence_of_element_located((By.ID, "password-input"))
    )
    pwd_input.clear()
    pwd_input.send_keys(PASSWORD)
    pwd_input.send_keys(Keys.ENTER)

    WebDriverWait(driver, LONG_WAIT).until(EC.url_contains(LCR_BASE))
    log("Login submitted successfully")


def build_requests_session_from_driver(driver: webdriver.Chrome) -> requests.Session:
    session = requests.Session()

    for cookie in driver.get_cookies():
        session.cookies.set(
            cookie["name"],
            cookie["value"],
            domain=cookie.get("domain"),
            path=cookie.get("path", "/"),
        )

    session.headers.update(
        {
            "User-Agent": driver.execute_script("return navigator.userAgent;") or "Mozilla/5.0",
            "Accept": "application/json, text/plain, */*",
            "Referer": ATTENDANCE_PAGE_URL,
            "Origin": LCR_BASE,
        }
    )
    return session


# ============================================================
# ROSTER FROM THE WORKING RENDERED MEMBER LIST
# ============================================================

def extract_uuid_from_row(row) -> str:
    candidate_attributes = (
        "data-uuid",
        "data-person-uuid",
        "data-personuuid",
        "data-member-uuid",
        "id",
    )

    for attr in candidate_attributes:
        value = (row.get_attribute(attr) or "").strip()
        match = re.search(
            r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}",
            value,
        )
        if match:
            return match.group(0)

    for link in row.find_elements(By.CSS_SELECTOR, "a[href]"):
        href = link.get_attribute("href") or ""
        match = re.search(
            r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}",
            href,
        )
        if match:
            return match.group(0)

    return ""


def fetch_member_roster_from_page(driver: webdriver.Chrome) -> tuple[Dict[str, str], Set[str]]:
    log(f"Loading rendered member list: {MEMBER_LIST_PAGE_URL}")
    driver.get(MEMBER_LIST_PAGE_URL)

    WebDriverWait(driver, ROSTER_WAIT).until(
        lambda d: (
            "Name" in get_body_text(d)
            and "Gender" in get_body_text(d)
            and "Birth Date" in get_body_text(d)
            and get_body_text(d).count(",") > 50
        )
    )

    DEBUG_MEMBER_HTML.write_text(driver.page_source, encoding="utf-8")
    DEBUG_MEMBER_TEXT.write_text(get_body_text(driver), encoding="utf-8")

    roster_by_uuid: Dict[str, str] = {}
    roster_names: Set[str] = set()
    debug_rows: list[str] = []

    for row in driver.find_elements(By.CSS_SELECTOR, "tr"):
        cells = row.find_elements(By.CSS_SELECTOR, "td")
        cell_texts = [cell.text.strip() for cell in cells]
        if cell_texts:
            debug_rows.append(" | ".join(cell_texts))

        if len(cell_texts) < 5:
            continue

        possible_name = clean_name(cell_texts[1])
        gender = cell_texts[2].strip()
        age = cell_texts[3].strip()
        birth_date = cell_texts[4].strip()

        if gender not in {"M", "F"}:
            continue
        if not age.isdigit():
            continue
        if not re.search(r"\b\d{1,2}\s+[A-Za-z]{3}\s+\d{4}\b", birth_date):
            continue
        if not looks_like_name(possible_name):
            continue

        roster_names.add(possible_name)
        uuid = extract_uuid_from_row(row)
        if uuid:
            roster_by_uuid[uuid] = possible_name

    DEBUG_MEMBER_ROWS.write_text("\n".join(debug_rows), encoding="utf-8")

    log(f"Rendered roster names found: {len(roster_names)}")
    log(f"Rendered roster UUID mappings found: {len(roster_by_uuid)}")

    if len(roster_names) < 50:
        raise RuntimeError(
            f"Only found {len(roster_names)} roster names. Member list may not have fully loaded."
        )

    return roster_by_uuid, roster_names


# ============================================================
# ATTENDANCE API
# ============================================================

def attendance_api_url(unit_number: str, start_dt: date, end_dt: date) -> str:
    return (
        f"{LCR_BASE}/api/umlu/v1/class-and-quorum/attendance/overview/"
        f"unitNumber/{unit_number}/start/{to_iso(start_dt)}/end/{to_iso(end_dt)}?lang=eng"
    )


def fetch_json(session: requests.Session, url: str) -> dict | list:
    response = session.get(url, timeout=60, allow_redirects=True)

    content_type = response.headers.get("content-type", "")
    if response.status_code >= 400:
        raise RuntimeError(
            f"Request failed with HTTP {response.status_code}: {url}\n"
            f"Response preview: {response.text[:500]}"
        )

    if "json" not in content_type.lower():
        raise RuntimeError(
            f"Expected JSON but received {content_type or 'unknown content type'} from {url}. "
            "The session may have been redirected to a login or error page.\n"
            f"Response preview: {response.text[:500]}"
        )

    return response.json()


def fetch_attendance_window(
    session: requests.Session,
    unit_number: str,
    start_dt: date,
    end_dt: date,
) -> dict:
    url = attendance_api_url(unit_number, start_dt, end_dt)
    log(f"Fetching attendance window: {to_iso(start_dt)} to {to_iso(end_dt)}")
    payload = fetch_json(session, url)

    if not isinstance(payload, dict):
        raise RuntimeError("Attendance API did not return an object.")

    DEBUG_ATTENDANCE_JSON.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    return payload


# ============================================================
# MERGE ATTENDANCE
# ============================================================

def name_from_attendance_person(person: dict) -> str:
    direct = first_string(
        person,
        (
            "nameListPreferredLocal",
            "listPreferredLocal",
            "displayName",
            "name",
            "memberName",
            "personName",
            "houseHoldMemberNameForList",
        ),
    )
    if direct:
        return clean_name(direct)

    name_formats = person.get("nameFormats")
    if isinstance(name_formats, dict):
        nested = first_string(
            name_formats,
            ("listPreferredLocal", "nameListPreferredLocal", "displayName", "name"),
        )
        if nested:
            return clean_name(nested)

    return ""


def merge_attendance_window(
    payload: dict,
    roster_by_uuid: Dict[str, str],
    roster_names: Set[str],
    attendance_data: Dict[str, Dict[date, bool]],
    all_dates: Set[date],
    start_dt: date,
    end_dt: date,
) -> tuple[int, int]:
    attendance_data_obj = payload.get("attendanceData") or payload
    attendees = attendance_data_obj.get("attendees") or []

    merged_people = 0
    skipped_people = 0

    for person in attendees:
        uuid = first_string(person, ("uuid", "personUuid", "personUUID", "memberUuid"))

        # Prefer the attendance response's own name. This removes dependence on
        # the member-list API that recently stopped working.
        name = name_from_attendance_person(person)
        if not name and uuid:
            name = roster_by_uuid.get(uuid, "")

        if not name or not looks_like_name(name):
            skipped_people += 1
            continue

        # Keep legitimate attendance names even if the rendered roster omits
        # an out-of-unit attendee, but log roster mismatches for diagnosis.
        if roster_names and name not in roster_names:
            log(f"Attendance name not found in rendered roster: {name}")

        merged_people += 1
        attendance_data.setdefault(name, {})

        entries = person.get("entries") or person.get("attendanceEntries") or []
        for entry in entries:
            raw_date = entry.get("date")
            if isinstance(raw_date, dict):
                iso = first_string(raw_date, ("isoYearMonthDay", "isoDate", "date"))
            elif isinstance(raw_date, str):
                iso = raw_date
            else:
                iso = first_string(entry, ("isoYearMonthDay", "attendanceDate"))

            if not iso:
                continue

            try:
                dt = datetime.strptime(iso[:10], "%Y-%m-%d").date()
            except ValueError:
                continue

            if not (start_dt <= dt <= end_dt):
                continue

            all_dates.add(dt)
            attended = bool(
                entry.get(
                    "isMarkedAttended",
                    entry.get("markedAttended", entry.get("attended", False)),
                )
            )
            attendance_data[name][dt] = attended

    return merged_people, skipped_people


def scrape_attendance_via_api(
    session: requests.Session,
    unit_number: str,
    start_dt: date,
    end_dt: date,
    roster_by_uuid: Dict[str, str],
    roster_names: Set[str],
) -> Tuple[Dict[str, Dict[date, bool]], List[date]]:
    attendance_data: Dict[str, Dict[date, bool]] = {}
    all_dates: Set[date] = set()

    windows = build_date_windows(start_dt, end_dt)
    log(f"Date windows to request: {[f'{to_iso(s)}..{to_iso(e)}' for s, e in windows]}")

    for win_start, win_end in windows:
        payload = fetch_attendance_window(session, unit_number, win_start, win_end)
        merged_people, skipped_people = merge_attendance_window(
            payload,
            roster_by_uuid,
            roster_names,
            attendance_data,
            all_dates,
            start_dt,
            end_dt,
        )
        log(f"Merged attendees from window: {merged_people}; skipped without usable names: {skipped_people}")

    final_dates = sorted(all_dates)
    if not final_dates:
        raise RuntimeError(
            "No attendance dates were collected from the API. "
            f"Inspect {DEBUG_ATTENDANCE_JSON}."
        )

    return attendance_data, final_dates


# ============================================================
# EXCEL OUTPUT
# ============================================================

def write_excel(
    attendance_data: Dict[str, Dict[date, bool]],
    all_dates: List[date],
    out_path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    percent_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    headers = ["Name", "% activity"] + [format_excel_header(dt) for dt in all_dates]
    ws.append(headers)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = left if col_idx == 1 else center

    for row_idx, name in enumerate(
        sorted(attendance_data.keys(), key=lambda s: s.casefold()), start=2
    ):
        per_date = attendance_data[name]
        total = len(all_dates)
        present_count = sum(1 for d in all_dates if per_date.get(d, False))
        pct = (present_count / total) if total else 0.0

        ws.cell(row=row_idx, column=1, value=name)

        pct_cell = ws.cell(row=row_idx, column=2, value=pct)
        pct_cell.number_format = "0%"
        pct_cell.fill = percent_fill

        for col_idx, dt in enumerate(all_dates, start=3):
            ws.cell(
                row=row_idx,
                column=col_idx,
                value="☑" if per_date.get(dt, False) else "☐",
            )

    ws.freeze_panes = "C2"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12

    for col_idx in range(3, 3 + len(all_dates)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.alignment = left if cell.column == 1 else center

    wb.save(out_path)


# ============================================================
# MAIN
# ============================================================

def main() -> int:
    if not USERNAME or not PASSWORD:
        err("Missing LCR_USERNAME and/or LCR_PASSWORD environment variables.")
        return 1

    start_dt = parse_iso_date(START_DATE)
    end_dt = parse_iso_date(END_DATE)

    if start_dt > end_dt:
        err("START_DATE must be on or before END_DATE.")
        return 1

    output_dir = ensure_dir(OUTPUT_DIR)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = output_dir / (
        f"attendance_{start_dt.isoformat()}_to_{end_dt.isoformat()}_{timestamp}.xlsx"
    )

    driver = make_driver()
    try:
        login(driver)

        # FIX: Load and parse the same rendered member-list page used by the
        # repaired fetch_lcr_all_names.py rather than calling the obsolete API.
        roster_by_uuid, roster_names = fetch_member_roster_from_page(driver)

        # Open the attendance page before copying cookies so all LCR session
        # cookies and attendance-specific context are present.
        log(f"Loading attendance page: {ATTENDANCE_PAGE_URL}")
        driver.get(ATTENDANCE_PAGE_URL)
        WebDriverWait(driver, LONG_WAIT).until(
            lambda d: "attendance" in d.current_url.lower()
            or "Attendance" in get_body_text(d)
        )

        session = build_requests_session_from_driver(driver)
        attendance_data, all_dates = scrape_attendance_via_api(
            session,
            UNIT_NUMBER,
            start_dt,
            end_dt,
            roster_by_uuid,
            roster_names,
        )

        write_excel(attendance_data, all_dates, out_path)
        log(f"Excel output written to {out_path}")
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    return 0


if __name__ == "__main__":
    sys.exit(main())
